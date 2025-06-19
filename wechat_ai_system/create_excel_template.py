import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = '库存模板'

headers = ['产品ID', '产品名称', '库存数量', '单价', '单位', '产品描述', '生产日期', '保质期']

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 10

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

sample_data = [
    ['001', '红茶特级', 100, 68.5, '盒', '特级红茶，口感醇厚', '2025-01-15', '24个月'],
    ['002', '绿茶珍品', 80, 45.0, '盒', '珍品绿茶，清香怡人', '2025-02-20', '18个月'],
    ['003', '乌龙茶', 120, 88.0, '袋', '优质乌龙茶，香气持久', '2025-03-10', '24个月'],
    ['004', '普洱茶饼', 50, 128.5, '饼', '陈年普洱，回甘悠长', '2024-12-05', '36个月'],
    ['005', '茉莉花茶', 90, 35.0, '盒', '茉莉花茶，花香四溢', '2025-01-30', '12个月']
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.border = thin_border
        if col_num in [3, 4]:  # Numeric columns
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

wb.save('inventory_template.xlsx')
print('Excel模板已创建: inventory_template.xlsx')
